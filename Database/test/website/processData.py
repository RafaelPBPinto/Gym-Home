import sqlite3
import openpyxl
import os
def processData():
    conn = sqlite3.connect('..\PlanosUser.db')
    conn.execute('PRAGMA foreign_keys = ON')
    listaExer = 'exercicio.xlsx'
    plaExer = 'DocemntosExecel\ExercicioPlanos.xlsx'
    plano = 'DocemntosExecel\Planos.xlsx'
    sessao = 'DocemntosExecel\Sessao.xlsx'
    if os.path.exists(listaExer):
            xlsx = openpyxl.load_workbook(listaExer)
            sheet = xlsx.active
            #dimensao = sheet.dimensions
            #values = sheet[dimensao]
            linha = sheet.rows
            for row in linha:
                if row[2].value != None:    # Eu acho que alguns erros de None acontecem porque se adiciona uma linha em branco no final do ficheiro excel
                    nome      = row[0].value
                    tipo      = row[1].value
                    duracao   = int(row[2].value)
                    descricao = row[3].value
                    if nome != None:
                        conn.execute('PRAGMA foreign_keys = ON')
                        query = f"INSERT INTO Exercicio (Nome,Tipo,Duracao,Descricao) VALUES ('{nome}', '{tipo}', '{duracao}', '{descricao}')"
                        conn.execute(query)
                        conn.commit()
    ###############################################################################            
    #Planos:
    if os.path.exists(plano):
        xlsx = openpyxl.load_workbook(plano)
        sheet = xlsx.active
        linha = sheet.rows
        for row in linha:

            nome      = row[0].value
            descricao = row[1].value
            autor     = row[2].value

            if nome != None:

                query = f"INSERT INTO Plano (Nome, Descricao, Autor) VALUES ('{nome}', '{descricao}', '{autor}')"
                conn.execute(query)
                conn.commit()

    #plano de exercicios: 
    if os.path.exists(plaExer):
            xlsx = openpyxl.load_workbook(plaExer)
            sheet = xlsx.active
            linha = sheet.rows
            for row in linha:

                tipo            = row[0].value
                series          = int(row[1].value)
                repeticao       = int(row[2].value)
                ordem           = int(row[4].value)
                RefID_exercicio = int(row[5].value)
                RefID_plano     = int(row[6].value)
                
                if tipo != None:
                   
                    query = f"INSERT INTO ExercicioPlano (Series,Repeticoes, Ordem,  RefID_plano, RefID_exercicio)\
                            VALUES ('{series}', '{repeticao}', '{ordem}', '{RefID_plano}', '{RefID_exercicio}')"
                    conn.execute(query)
                    conn.commit()

   
    
    #########################################################################################
    #Sessão:
    if os.path.exists(sessao):
        xlsx = openpyxl.load_workbook(sessao)
        sheet = xlsx.active
        linha = sheet.rows
        for row in linha:
            if row[0].value != None:
                RefID_utilizador  = int(row[0].value)
                RefID_plano       = int(row[1].value)
                Dia               = row[2].value
                if Dia != None:
                    query = f"INSERT INTO Sessao (RefID_utilizador, RefID_plano, Dia) VALUES ('{RefID_utilizador}', '{RefID_plano}', '{Dia}')"
                    conn.execute(query)
                    conn.commit()


    #######################################################################
    #img de exer1
    photo = convert2BinData('imagem\ponte.jpg') 
    insertImage("Ponte",photo,1)
    #img de exer2
    photo = convert2BinData('imagem\AgachamentoCadeira.jpg') 
    insertImage('Agachamento na cadeira',photo,2)
    #img de exer3
    photo = convert2BinData('imagem\Gemeos.jpg') 
    insertImage('Gemeos',photo,3)
    #img de exer4
    photo = convert2BinData('imagem\legExtension.jpg') 
    insertImage('Leg Extension',photo,4)
    #img de exer5
    photo = convert2BinData('imagem\FlexaoParede.jpg') 
    insertImage('Flexão contra a parede',photo,5)
    #img de exer6
    photo = convert2BinData('imagem\Abdominal.png')  
    insertImage('Abdominal',photo,6)
    #img de exer7
    photo = convert2BinData('imagem\singleLegStand.jpg') 
    insertImage('Apoio numa perna',photo,7)
    #img de exer8
    photo = convert2BinData('imagem\TandemStanding.jpg') 
    insertImage('Pés seguidos',photo,8)
    #img de exer9
    photo = convert2BinData('imagem\sideBendStretch.jpg') 
    insertImage('Alongamento curva lateral',photo,9)
    #img de exer10
    photo = convert2BinData('imagem\AlongamentoPeito.jpg') 
    insertImage('Alongamento peito',photo,10)
    #img de exer11
    photo = convert2BinData('imagem\TandemWalking.png')
    insertImage('Andar em pés seguidos',photo,11)
    #img de exer12
    photo = convert2BinData('imagem\ChutoLateral.jpg') 
    insertImage('Chuto lateral',photo,12)
    #img de exer13
    photo = convert2BinData('imagem\AlongamentoOmbro.jpg') 
    insertImage('Alongamento ombro',photo,13)
    #img de exer14
    photo = convert2BinData('imagem\AlongamentoQuadricep.jpg') 
    insertImage('Alongamento quadríceps',photo,14)
    #img de exer15
    photo = convert2BinData('imagem\AlongamentoAbdominal.jpg') 
    insertImage('Alongamento abdominal',photo,15)


def convert2BinData(filename):
    with open(filename, 'rb') as f:
        binData = f.read()
    return binData

def insertImage(nome,img,id):
    conn = sqlite3.connect('..\PlanosUser.db')
    conn.execute('PRAGMA foreign_keys = ON')
    #query = f"INSERT INTO Imagem (Nome,ImagemBinary,RefID_exercicio) VALUES ('{nome}','{img}','{id}')"
    query = """ INSERT INTO Imagem (Nome, ImagemBinary,RefID_exercicio) VALUES (?, ?, ?)"""
    data_tuple = (nome,img,id)
    conn.execute(query,data_tuple)
    conn.commit()
    conn.close()

if __name__ == '__main__':
    processData()